#!/usr/bin/env python3
"""
SQL Injection Vulnerability Scanner Core Module

Provides the SQLInjectionScanner class for scanning Python files and generating
vulnerability reports in Excel format.
"""

import os
import re
import sys
from pathlib import Path
from typing import List, Dict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError(
        "openpyxl is required. Install with: pip install openpyxl"
    )


class SQLInjectionScanner:
    """
    Scanner for detecting potential SQL injection vulnerabilities in Python code.
    
    This scanner identifies dangerous patterns such as:
    - f-strings in SQL queries
    - String concatenation in SQL
    - .format() method usage
    - % formatting in SQL
    
    Attributes:
        root_dir (Path): The root directory to scan
        findings (List[Dict]): List of discovered vulnerabilities
    """
    
    # Patterns that indicate potential SQL injection vulnerabilities
    VULNERABLE_PATTERNS = [
        (r'f["\'].*?(?:SELECT|INSERT|UPDATE\s+|DELETE\s+FROM|DROP\s+|CREATE\s+|ALTER\s+|EXEC|EXECUTE).*?\{.*?\}.*?["\']', 'f-string with SQL'),
        (r'["\'].*?(?:SELECT|INSERT|UPDATE\s+|DELETE\s+FROM|DROP\s+|CREATE\s+|ALTER\s+).*?["\']\s*\+\s*', 'String concatenation in SQL'),
        (r'["\'].*?(?:SELECT|INSERT|UPDATE\s+|DELETE\s+FROM|DROP\s+|CREATE\s+|ALTER\s+).*?["\']\.format\(', '.format() with SQL'),
        (r'["\'].*?(?:SELECT|INSERT|UPDATE\s+|DELETE\s+FROM|DROP\s+|CREATE\s+|ALTER\s+).*?["\']\s*%\s*', '% formatting in SQL'),
    ]
    
    # Patterns that indicate safe parameterized queries
    SAFE_PATTERNS = [
        r'execute\s*\(\s*["\'][^"\']*["\'],\s*\[',
        r'execute\s*\(\s*["\'][^"\']*["\'],\s*\(',
        r'%s.*?[,)]',
        r'\?',
    ]
    
    # SQL keywords to look for
    SQL_KEYWORDS = {'SELECT', 'INSERT', 'UPDATE', 'DELETE', 'DROP', 'CREATE', 'ALTER', 'EXECUTE', 'EXEC'}
    
    # Directories to skip during scanning
    SKIP_DIRS = {'.git', '__pycache__', 'node_modules', '.venv', 'venv', '.env', 'dist', 'build', '*.egg-info'}

    def __init__(self, root_dir: str = '.'):
        """
        Initialize the SQL injection scanner.
        
        Args:
            root_dir (str): The root directory to scan. Defaults to current directory.
        """
        self.root_dir = Path(root_dir)
        self.findings: List[Dict] = []

    def find_python_files(self) -> List[Path]:
        """
        Find all Python files in the root directory.
        
        Returns:
            List[Path]: List of Python file paths.
        """
        python_files = []
        for root, dirs, files in os.walk(self.root_dir):
            # Skip specified directories
            dirs[:] = [d for d in dirs if d not in self.SKIP_DIRS]
            
            for file in files:
                if file.endswith('.py'):
                    python_files.append(Path(root) / file)
        
        return python_files

    def is_safe_query(self, line: str) -> bool:
        """
        Check if a line uses safe parameterized queries.
        
        Args:
            line (str): The line of code to check.
            
        Returns:
            bool: True if the line appears to use safe parameterized queries.
        """
        for pattern in self.SAFE_PATTERNS:
            if re.search(pattern, line):
                return True
        return False

    def _is_logging_or_print_statement(self, line: str) -> bool:
        """
        Check if a line is a logging or print statement.
        
        Args:
            line (str): The line of code to check.
            
        Returns:
            bool: True if the line is a logging or print statement.
        """
        logging_patterns = [
            r'print\s*\(',
            r'\blog\s*\.',
            r'\blogger\s*\.',
            r'\blogging\s*\.',
            r'\.info\s*\(',
            r'\.debug\s*\(',
            r'\.error\s*\(',
            r'\.warning\s*\(',
            r'\.exception\s*\(',
            r'\.trace\s*\(',
            r'"error"\s*:',
            r"'error'\s*:",
        ]
        for pattern in logging_patterns:
            if re.search(pattern, line, re.IGNORECASE):
                return True
        return False

    def _is_likely_sql_construction(self, line: str) -> bool:
        """
        Check if a line looks like SQL query construction or execution.
        
        Args:
            line (str): The line of code to check.
            
        Returns:
            bool: True if the line appears to be constructing or executing a SQL query.
        """
        sql_construction_patterns = [
            r'\bquery\s*=',
            r'\bsql\s*=',
            r'\.execute\s*\(',
            r'\.executemany\s*\(',
            r'\.query\s*\(',
            r'\bcursor\s*\.',
            r'\bdb\s*\.',
            r'\bdatabase\s*\.',
            r'\bconn\s*\.',
            r'\bconnection\s*\.',
        ]
        for pattern in sql_construction_patterns:
            if re.search(pattern, line, re.IGNORECASE):
                return True
        return False

    def _calculate_risk_level(self, line: str) -> str:
        """
        Calculate the risk level of a potential SQL injection vulnerability.
        
        Args:
            line (str): The line of code to analyze.
            
        Returns:
            str: Risk level ('HIGH', 'MEDIUM', or 'LOW').
        """
        if 'f"' in line or "f'" in line:
            # Check if it's actually a SQL context (contains SQL keywords with word boundaries)
            if re.search(r'\b(SELECT|INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|EXEC|EXECUTE)\b', line, re.IGNORECASE):
                return 'HIGH'
        elif '+' in line and re.search(r'\b(SELECT|INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|EXEC|EXECUTE)\b', line, re.IGNORECASE):
            return 'HIGH'
        elif '.format(' in line and re.search(r'\b(SELECT|INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|EXEC|EXECUTE)\b', line, re.IGNORECASE):
            return 'MEDIUM'
        elif '%' in line and re.search(r'\b(SELECT|INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|EXEC|EXECUTE)\b', line, re.IGNORECASE):
            return 'MEDIUM'
        
        return 'LOW'

    def detect_sql_injection(self, file_path: Path, content: str) -> List[Dict]:
        """
        Detect potential SQL injection vulnerabilities in a file.
        
        Args:
            file_path (Path): Path to the file being scanned.
            content (str): The file content.
            
        Returns:
            List[Dict]: List of vulnerability findings.
        """
        findings = []
        lines = content.split('\n')
        
        for line_num, line in enumerate(lines, 1):
            # Skip comments
            if line.strip().startswith('#'):
                continue
            
            # Skip logging and print statements
            if self._is_logging_or_print_statement(line):
                continue
            
            # Check if line contains SQL keywords
            if not any(keyword in line.upper() for keyword in self.SQL_KEYWORDS):
                continue
            
            # Check if this line looks like SQL construction/execution
            if not self._is_likely_sql_construction(line):
                continue
            
            # Skip if it's a safe parameterized query
            if self.is_safe_query(line):
                continue
            
            # Check for SQL injection patterns
            for pattern, description in self.VULNERABLE_PATTERNS:
                if re.search(pattern, line, re.IGNORECASE):
                    risk_level = self._calculate_risk_level(line)
                    findings.append({
                        'file_path': str(file_path.relative_to(self.root_dir)),
                        'module_name': file_path.stem,
                        'line_number': line_num,
                        'code': line.strip(),
                        'pattern': description,
                        'risk_level': risk_level
                    })
                    break
        
        return findings

    def scan(self) -> List[Dict]:
        """
        Scan all Python files for SQL injection vulnerabilities.
        
        Returns:
            List[Dict]: List of discovered vulnerabilities.
        """
        python_files = self.find_python_files()
        
        print(f"Scanning {len(python_files)} Python files...")
        
        for file_path in python_files:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    findings = self.detect_sql_injection(file_path, content)
                    self.findings.extend(findings)
            except Exception as e:
                print(f"Warning: Error scanning {file_path}: {e}")
        
        return self.findings

    def generate_excel_report(self, output_file: str = 'sql_injection_report.xlsx') -> str:
        """
        Generate an Excel report of all findings.
        
        Args:
            output_file (str): Path to the output Excel file.
            
        Returns:
            str: Path to the generated report file.
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SQL Injection Findings"
        
        # Define styles
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        low_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['File Path', 'Module Name', 'Line Number', 'Pattern', 'Risk Level', 'Code Snippet']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Sort findings by risk level (HIGH first) and then by file path
        risk_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
        sorted_findings = sorted(
            self.findings,
            key=lambda x: (risk_order.get(x['risk_level'], 3), x['file_path'], x['line_number'])
        )
        
        # Add data rows
        for finding in sorted_findings:
            ws.append([
                finding['file_path'],
                finding['module_name'],
                finding['line_number'],
                finding['pattern'],
                finding['risk_level'],
                finding['code']
            ])
            
            row = ws.max_row
            # Apply risk level color
            risk_fill = {
                'HIGH': high_fill,
                'MEDIUM': medium_fill,
                'LOW': low_fill
            }.get(finding['risk_level'], PatternFill())
            
            for col in range(1, 7):
                cell = ws.cell(row, col)
                cell.fill = risk_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 50
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(['SQL Injection Vulnerability Report Summary'])
        summary_ws.append([])
        summary_ws.append(['Total Findings', len(self.findings)])
        
        high_count = sum(1 for f in self.findings if f['risk_level'] == 'HIGH')
        medium_count = sum(1 for f in self.findings if f['risk_level'] == 'MEDIUM')
        low_count = sum(1 for f in self.findings if f['risk_level'] == 'LOW')
        
        summary_ws.append(['High Risk', high_count])
        summary_ws.append(['Medium Risk', medium_count])
        summary_ws.append(['Low Risk', low_count])
        summary_ws.append([])
        summary_ws.append(['Report generated from:', str(self.root_dir)])
        
        for row in summary_ws.iter_rows():
            for cell in row:
                cell.border = border
        
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 15
        
        # Save workbook
        wb.save(output_file)
        return output_file
